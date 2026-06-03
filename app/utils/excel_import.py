"""
Excel批量导入工具模块

支持从Excel文件读取数据、字段校验、重复检查、批量插入数据库。
所有模块的导入逻辑统一在此实现，保持代码复用和一致性。
"""
import os
import re
import traceback
from datetime import datetime, date
from io import BytesIO
from collections import OrderedDict

import openpyxl
from flask import current_app


# ============================================================
# 通用工具函数
# ============================================================

def allowed_file(filename, extensions=None):
    """检查文件扩展名是否允许上传"""
    if extensions is None:
        extensions = {'xlsx', 'xls'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in extensions


def parse_excel(file_storage, sheet_index=0, header_row=1):
    """
    解析Excel文件，返回列名列表和数据行列表

    :param file_storage: Flask上传文件对象
    :param sheet_index: 工作表索引（从0开始）
    :param header_row: 表头所在行（从1开始）
    :return: (columns, rows) 列名列表, 数据行(字典)列表
    """
    wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if sheet_index >= len(sheet_names):
        raise ValueError(f'工作表索引 {sheet_index} 超出范围，文件仅有 {len(sheet_names)} 个工作表')

    ws = wb[sheet_names[sheet_index]]

    rows_iter = ws.iter_rows(values_only=True)
    all_rows = list(rows_iter)
    wb.close()

    if not all_rows:
        return [], []

    # 提取表头
    header_index = header_row - 1
    if header_index >= len(all_rows):
        return [], []

    raw_headers = all_rows[header_index]
    columns = []
    for h in raw_headers:
        if h is None:
            columns.append('')
        else:
            col = str(h).strip()
            columns.append(col)

    # 提取数据行
    data_rows = []
    for row in all_rows[header_index + 1:]:
        # 跳过全空行
        if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
            continue
        row_dict = OrderedDict()
        for idx, col_name in enumerate(columns):
            if col_name:
                value = row[idx] if idx < len(row) else None
                row_dict[col_name] = value
        data_rows.append(row_dict)

    return columns, data_rows


def safe_str(value, default=''):
    """安全转换为字符串"""
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def safe_int(value, default=None):
    """安全转换为整数"""
    if value is None:
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def safe_float(value, default=None):
    """安全转换为浮点数"""
    if value is None:
        return default
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return default


def safe_date(value, default=None):
    """安全转换为日期对象"""
    if value is None:
        return default
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        # 尝试常见格式
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日',
                     '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
    except Exception:
        pass
    return default


def safe_datetime(value, default=None):
    """安全转换为日期时间对象"""
    if value is None:
        return default
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    try:
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                     '%Y-%m-%d', '%Y/%m/%d'):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
    except Exception:
        pass
    return default


# ============================================================
# 字段校验器
# ============================================================

class FieldValidator:
    """
    字段校验器，支持必填检查、类型转换、值范围校验等
    """

    def __init__(self, field_name, display_name=None, required=False,
                 field_type='str', min_value=None, max_value=None,
                 max_length=None, regex=None, options=None,
                 default=None, unique_check=None):
        """
        :param field_name: Excel列名（也是数据库字段名）
        :param display_name: 显示名称（用于错误提示）
        :param required: 是否必填
        :param field_type: 字段类型: str, int, float, date, datetime
        :param min_value: 最小值（int/float类型）
        :param max_value: 最大值（int/float类型）
        :param max_length: 最大长度（str类型）
        :param regex: 正则表达式校验
        :param options: 可选值列表
        :param default: 默认值
        :param unique_check: 唯一性检查函数, 接收(value) -> bool
        """
        self.field_name = field_name
        self.display_name = display_name or field_name
        self.required = required
        self.field_type = field_type
        self.min_value = min_value
        self.max_value = max_value
        self.max_length = max_length
        self.regex = regex
        self.options = options
        self.default = default
        self.unique_check = unique_check

    def validate(self, value, row_index):
        """
        校验单个字段值

        :param value: 原始值
        :param row_index: 行号（用于错误提示）
        :return: (cleaned_value, error_message)
        """
        # 空值处理
        if value is None or (isinstance(value, str) and value.strip() == ''):
            if self.required:
                return None, f'第{row_index}行「{self.display_name}」不能为空'
            return self.default, None

        # 类型转换
        cleaned = None
        try:
            if self.field_type == 'int':
                cleaned = safe_int(value)
                if cleaned is None:
                    return None, f'第{row_index}行「{self.display_name}」必须是整数'
            elif self.field_type == 'float':
                cleaned = safe_float(value)
                if cleaned is None:
                    return None, f'第{row_index}行「{self.display_name}」必须是数字'
            elif self.field_type == 'date':
                cleaned = safe_date(value)
                if cleaned is None:
                    return None, f'第{row_index}行「{self.display_name}」日期格式无效（正确格式：YYYY-MM-DD）'
            elif self.field_type == 'datetime':
                cleaned = safe_datetime(value)
                if cleaned is None:
                    return None, f'第{row_index}行「{self.display_name}」日期时间格式无效'
            else:
                cleaned = safe_str(value)
        except Exception:
            return None, f'第{row_index}行「{self.display_name}」格式错误'

        # 范围校验
        if self.field_type in ('int', 'float') and cleaned is not None:
            if self.min_value is not None and cleaned < self.min_value:
                return None, f'第{row_index}行「{self.display_name}」不能小于{self.min_value}'
            if self.max_value is not None and cleaned > self.max_value:
                return None, f'第{row_index}行「{self.display_name}」不能大于{self.max_value}'

        # 字符串长度校验
        if self.field_type == 'str' and self.max_length is not None and len(cleaned) > self.max_length:
            return None, f'第{row_index}行「{self.display_name}」长度不能超过{self.max_length}个字符'

        # 正则校验
        if self.regex and not re.match(self.regex, str(cleaned)):
            return None, f'第{row_index}行「{self.display_name}」格式不正确'

        # 选项校验
        if self.options and str(cleaned) not in [str(o) for o in self.options]:
            return None, f'第{row_index}行「{self.display_name}」值不在可选范围内: {", ".join(str(o) for o in self.options)}'

        # 唯一性检查
        if self.unique_check and cleaned is not None:
            if self.unique_check(cleaned):
                return None, f'第{row_index}行「{self.display_name}」值"{cleaned}"已存在'

        return cleaned, None


# ============================================================
# Excel导入器
# ============================================================

class ExcelImporter:
    """
    Excel批量导入器

    使用方式:
        importer = ExcelImporter(file, validators=[...])
        result = importer.run(process_func=my_func)

    其中 process_func(row_data, row_index) -> (success, error_msg)
    """

    def __init__(self, file_storage, validators=None, sheet_index=0, header_row=1):
        """
        :param file_storage: Flask上传文件对象
        :param validators: FieldValidator列表
        :param sheet_index: 工作表索引
        :param header_row: 表头行
        """
        self.file_storage = file_storage
        self.validators = validators or []
        self.sheet_index = sheet_index
        self.header_row = header_row
        self.columns = []
        self.rows = []

    def parse(self):
        """解析Excel文件"""
        self.columns, self.rows = parse_excel(
            self.file_storage,
            sheet_index=self.sheet_index,
            header_row=self.header_row
        )
        return len(self.rows)

    def validate_all(self):
        """
        校验所有数据行

        :return: (is_valid, errors)
            is_valid: True表示全部校验通过
            errors: 错误信息列表，每项为 {row: int, field: str, message: str}
        """
        errors = []
        for idx, row in enumerate(self.rows):
            row_index = self.header_row + 1 + idx  # Excel中的实际行号
            for validator in self.validators:
                raw_value = row.get(validator.field_name)
                cleaned, error = validator.validate(raw_value, row_index)
                if error:
                    errors.append({
                        'row': row_index,
                        'field': validator.display_name,
                        'message': error
                    })
                else:
                    # 将清洗后的值写回
                    row[validator.field_name] = cleaned

        return len(errors) == 0, errors

    def run(self, process_func, batch_size=100):
        """
        执行导入

        :param process_func: 处理函数, 接收(row_data, row_index) -> (success, error_msg)
        :param batch_size: 每批提交的行数
        :return: dict {success: int, fail: int, errors: list, total: int}
        """
        from app.models.base import db

        # 1. 解析
        try:
            total_rows = self.parse()
        except Exception as e:
            return {
                'success': 0,
                'fail': 0,
                'total': 0,
                'errors': [{'row': 0, 'field': '文件', 'message': f'文件解析失败: {str(e)}'}]
            }

        if total_rows == 0:
            return {'success': 0, 'fail': 0, 'total': 0,
                    'errors': [{'row': 0, 'field': '文件', 'message': 'Excel文件为空或没有数据行'}]
                    }

        # 2. 校验
        is_valid, validation_errors = self.validate_all()
        if not is_valid:
            return {
                'success': 0,
                'fail': len(self.rows),
                'total': len(self.rows),
                'errors': validation_errors
            }

        # 3. 逐行处理
        success_count = 0
        fail_count = 0
        all_errors = []
        batch_rows = []

        for idx, row in enumerate(self.rows):
            row_index = self.header_row + 1 + idx
            try:
                success, error_msg = process_func(row, row_index)
                if success:
                    success_count += 1
                    batch_rows.append(row)
                else:
                    fail_count += 1
                    all_errors.append({
                        'row': row_index,
                        'field': '数据处理',
                        'message': error_msg
                    })
            except Exception as e:
                fail_count += 1
                all_errors.append({
                    'row': row_index,
                    'field': '系统错误',
                    'message': f'处理异常: {str(e)}'
                })
                traceback.print_exc()

        return {
            'success': success_count,
            'fail': fail_count,
            'total': len(self.rows),
            'errors': all_errors
        }


# ============================================================
# 构建导入结果JSON响应
# ============================================================

def build_import_response(result):
    """
    根据导入结果构建统一的JSON响应

    :param result: run()方法返回的结果dict
    :return: Flask jsonify-ready dict
    """
    if result['fail'] > 0:
        # 有失败的情况
        error_preview = result['errors'][:10]  # 最多显示前10条错误
        return {
            'code': 200,
            'message': f'导入完成：成功 {result["success"]} 条，失败 {result["fail"]} 条',
            'data': {
                'success': result['success'],
                'fail': result['fail'],
                'total': result['total'],
                'errors': error_preview,
                'has_more_errors': len(result['errors']) > 10
            }
        }

    return {
        'code': 200,
        'message': f'成功导入 {result["success"]} 条数据',
        'data': {
            'success': result['success'],
            'fail': 0,
            'total': result['total'],
            'errors': []
        }
    }


def get_import_template(fields, sheet_name='Sheet1'):
    """
    生成导入模板Excel文件

    :param fields: 字段列表，每项为 {name, display_name, required, example}
    :param sheet_name: 工作表名称
    :return: BytesIO对象（可直接作为文件响应返回）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    headers = []
    for f in fields:
        header = f['display_name']
        if f.get('required'):
            header += ' *'
        headers.append(header)
    ws.append(headers)

    # 示例行（如果有）
    if any(f.get('example') is not None for f in fields):
        example_row = []
        for f in fields:
            example_row.append(f.get('example', ''))
        ws.append(example_row)

    # 设置列宽
    for col_idx, f in enumerate(fields, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        # 根据字段名和示例长度估算列宽
        estimated_width = max(len(f.get('display_name', '')), 10)
        if f.get('example'):
            estimated_width = max(estimated_width, len(str(f.get('example', ''))))
        ws.column_dimensions[col_letter].width = min(estimated_width + 4, 40)

    # 表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
